import streamlit as st
from ultralytics import YOLO
import cv2
import numpy as np
import time
import mysql.connector
import pandas as pd
from datetime import datetime
import plotly.graph_objects as go
import json
import os

# --- KONFIGURASI HALAMAN ---
st.set_page_config(page_title="Smart Counter Koperasi Merah Putih", layout="wide", initial_sidebar_state="expanded")

# --- MANAJEMEN KONFIGURASI (PERSISTENSI) ---
CONFIG_FILE = "ui_config.json"

def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            return json.load(f)
    return {
        "LINE_ORIENT": "Vertikal", "LINE_POS": 320, "ENTRY_DIR": "Kiri ke Kanan",
        "STAFF_LIMIT": 10, "BUYER_LIMIT": 10,
        "stf_x": 50, "stf_y": 50, "stf_w": 200, "stf_h": 300,
        "ksr_x": 380, "ksr_y": 50, "ksr_w": 200, "ksr_h": 300
    }

def save_config(config):
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f)

cfg = load_config()

# --- INJEKSI CSS UNTUK DESAIN DARK THEME SaaS ---
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {background-color: transparent !important;}
    
    .top-nav {
        display: flex; justify-content: space-between; align-items: center;
        background-color: #0A1628; padding: 10px 20px; border-bottom: 1px solid #1E4D8C;
        margin-top: -60px; margin-bottom: 20px; color: white; font-weight: 600;
        border-radius: 8px;
    }
    .status-live { color: #00C9A7; display: flex; align-items: center; font-size: 14px;}
    .status-live span { height: 10px; width: 10px; background-color: #00C9A7; border-radius: 50%; display: inline-block; margin-right: 8px; box-shadow: 0 0 8px #00C9A7;}
    
    .metric-card {
        background-color: #0F2040; border-left: 4px solid #1E4D8C;
        border-radius: 12px; padding: 16px; margin-bottom: 20px;
    }
    .metric-title { color: #A0AEC0; font-size: 12px; font-weight: 600; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 10px; }
    .metric-value { font-size: 36px; font-weight: 700; margin-bottom: 5px; }
    .val-white { color: #FFFFFF; }
    .val-teal { color: #00C9A7; }
    .val-pink { color: #D9568B; }
    .metric-sub { color: #718096; font-size: 12px; }
    
    .log-table { width: 100%; border-collapse: collapse; font-size: 13px; color: #A0AEC0;}
    .log-table th { text-align: left; padding: 12px 8px; border-bottom: 1px solid #1E4D8C; color: white;}
    .log-table td { padding: 10px 8px; border-bottom: 1px solid #0A1628;}
    .log-table tr:nth-child(even) { background-color: #0A1628; }
    .log-table tr:nth-child(odd) { background-color: #0F2040; }
    .badge-entry { background-color: rgba(0, 201, 167, 0.2); color: #00C9A7; padding: 4px 8px; border-radius: 4px; font-size: 11px;}
    .badge-buyer { background-color: rgba(217, 86, 139, 0.2); color: #D9568B; padding: 4px 8px; border-radius: 4px; font-size: 11px;}
    .badge-staff { background-color: rgba(30, 77, 140, 0.4); color: #82B1FF; padding: 4px 8px; border-radius: 4px; font-size: 11px;}
    
    .streamlit-expanderHeader { font-weight: 600 !important; color: white !important; }
    </style>
""", unsafe_allow_html=True)

# --- HEADER CUSTOM ---
st.markdown("""
<div class="top-nav">
    <div style="font-size: 18px; letter-spacing: 0.5px;">Smart Counter Koperasi Merah Putih</div>
    <div style="display: flex; gap: 20px; align-items: center; color: #A0AEC0; font-weight: 400; font-size: 14px;">
        <div class="status-live"><span></span> LIVE</div>
    </div>
</div>
""", unsafe_allow_html=True)

# --- KONEKSI DATABASE ---
@st.cache_resource
def init_connection():
    try:
        return mysql.connector.connect(
            host="localhost", user="root", password="", database="koperasi_db"
        )
    except Exception:
        return None

db = init_connection()
cursor = db.cursor() if db else None

def fetch_today_stats():
    if not cursor: return 0, 0, 0
    today = datetime.now().strftime('%Y-%m-%d')
    cursor.execute("SELECT COUNT(*) FROM visitor_logs WHERE event_type = 'Masuk' AND DATE(created_at) = %s", (today,))
    tin = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM visitor_logs WHERE event_type = 'Keluar' AND DATE(created_at) = %s", (today,))
    tout = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM visitor_logs WHERE event_type = 'Pembeli Baru' AND DATE(created_at) = %s", (today,))
    tbuy = cursor.fetchone()[0]
    return tin, tout, tbuy

def get_chart_data():
    if not cursor: return pd.DataFrame(), pd.DataFrame()
    today = datetime.now().strftime('%Y-%m-%d')
    query_hour = f"SELECT HOUR(created_at) as jam, COUNT(*) as total FROM visitor_logs WHERE event_type = 'Masuk' AND DATE(created_at) = '{today}' GROUP BY jam"
    df_hour = pd.read_sql(query_hour, db)
    
    query_days = """
        SELECT DATE(created_at) as tanggal, 
        SUM(CASE WHEN event_type = 'Masuk' THEN 1 ELSE 0 END) as Pengunjung,
        SUM(CASE WHEN event_type = 'Pembeli Baru' THEN 1 ELSE 0 END) as Pembeli
        FROM visitor_logs 
        WHERE created_at >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
        GROUP BY tanggal ORDER BY tanggal ASC
    """
    df_days = pd.read_sql(query_days, db)
    return df_hour, df_days

# FUNGSI BARU UNTUK EKSPOR CSV
def get_export_data(start_date, end_date):
    if not cursor: return pd.DataFrame()
    query = """
        SELECT 
            DATE(created_at) as Tanggal,
            SUM(CASE WHEN event_type = 'Masuk' THEN 1 ELSE 0 END) as 'Total Pengunjung',
            SUM(CASE WHEN event_type = 'Pembeli Baru' THEN 1 ELSE 0 END) as 'Total Pembeli',
            SUM(CASE WHEN event_type = 'Staf Aktif' THEN 1 ELSE 0 END) as 'Jumlah Staf'
        FROM visitor_logs
        WHERE DATE(created_at) >= %s AND DATE(created_at) <= %s
        GROUP BY Tanggal
        ORDER BY Tanggal ASC
    """
    cursor.execute(query, (start_date, end_date))
    rows = cursor.fetchall()
    cols = ['Tanggal', 'Total Pengunjung', 'Total Pembeli', 'Jumlah Staf']
    df = pd.DataFrame(rows, columns=cols)
    if not df.empty:
        df['Tingkat Konversi (%)'] = (df['Total Pembeli'] / df['Total Pengunjung'] * 100).fillna(0).round(1)
        df['Tingkat Konversi (%)'] = df['Tingkat Konversi (%)'].replace([np.inf, -np.inf], 0)
    return df

def log_to_database(track_id, event_type):
    if cursor:
        try:
            cursor.execute("INSERT INTO visitor_logs (track_id, event_type, created_at) VALUES (%s, %s, %s)", 
                           (int(track_id), event_type, datetime.now()))
            db.commit()
        except: pass

def add_log(track_id, event, zone, duration="-"):
    time_str = datetime.now().strftime("%H:%M:%S")
    st.session_state.recent_logs.insert(0, {"time": time_str, "id": f"#{track_id:04d}", "event": event, "zone": zone, "duration": duration})
    if len(st.session_state.recent_logs) > 6: st.session_state.recent_logs.pop()

def render_log_table():
    html = '<table class="log-table"><tr><th>Waktu</th><th>ID</th><th>Peristiwa</th><th>Zona</th><th>Durasi</th></tr>'
    for log in st.session_state.recent_logs:
        badge_class = "badge-entry" if log['event'] == "Masuk" else ("badge-buyer" if log['event'] == "Pembeli Baru" else "badge-staff")
        html += f"<tr><td>{log['time']}</td><td>{log['id']}</td><td><span class='{badge_class}'>{log['event']}</span></td><td>{log['zone']}</td><td>{log['duration']}</td></tr>"
    html += '</table>'
    return html

# --- INISIALISASI SESI & RESET HARIAN ---
today_str = datetime.now().strftime('%Y-%m-%d')
if 'initialized' not in st.session_state or st.session_state.get('current_date') != today_str:
    tin, tout, tbuy = fetch_today_stats()
    st.session_state.update({
        'initialized': True,
        'current_date': today_str,
        'count_in': tin, 'count_out': tout, 'count_buyer': tbuy,
        'track_states': {}, 'staff_zone_timers': {}, 'cashier_zone_timers': {},
        'staff_ids': set(), 'buyer_ids': set(), 'recent_logs': []
    })

# --- SIDEBAR CONTROLS ---
st.sidebar.markdown('<div style="color:white; font-size:18px; font-weight:600; margin-bottom:20px;">Panel Admin<br><span style="color:#A0AEC0; font-size:12px; font-weight:400;">Tampilan Operasional</span></div>', unsafe_allow_html=True)
video_source = st.sidebar.radio("Sumber Deteksi:", ["Kamera Live", "Unggah Video (MP4)"])

video_path = None
if video_source == "Kamera Live":
    cam_index = st.sidebar.number_input("Pilih ID Kamera", min_value=0, max_value=5, value=0)
    video_path = cam_index
else:
    uploaded_file = st.sidebar.file_uploader("Unggah File Video", type=['mp4', 'avi', 'mov'])
    if uploaded_file is not None:
        with open("temp_video.mp4", "wb") as f:
            f.write(uploaded_file.read())
        video_path = "temp_video.mp4"

run_camera = st.sidebar.toggle("Mulai Deteksi", value=False)

# === FITUR EKSPOR LAPORAN EXCEL ===
st.sidebar.markdown("---")
st.sidebar.markdown('<div style="color:white; font-size:16px; font-weight:600; margin-bottom:10px;">Ekspor Laporan</div>', unsafe_allow_html=True)
date_selection = st.sidebar.date_input("Pilih Rentang Tanggal:", value=[])

show_preview = False
df_export = pd.DataFrame()

if len(date_selection) > 0:
    start_date = date_selection[0]
    end_date = date_selection[1] if len(date_selection) > 1 else date_selection[0]
    
    df_export = get_export_data(start_date, end_date)
    
    if not df_export.empty:
        # 1. Menyisipkan kolom "No." di urutan paling kiri (indeks 0)
        df_export.insert(0, 'No.', range(1, len(df_export) + 1))
        
        # 2. Mengonversi DataFrame ke format Excel murni (.xlsx) beserta Styling-nya
        from io import BytesIO
        from openpyxl.styles import Border, Side, Font, Alignment
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Laporan KMP')
            
            worksheet = writer.sheets['Laporan KMP']
            
            # Membuat gaya garis batas (border tipis)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Menyesuaikan lebar kolom secara otomatis agar tanggal tidak menjadi ########
            for idx, col in enumerate(df_export.columns):
                col_letter = worksheet.cell(row=1, column=idx+1).column_letter
                # Menghitung panjang teks maksimal di setiap kolom
                max_len = max(df_export[col].astype(str).map(len).max(), len(str(col))) + 4
                worksheet.column_dimensions[col_letter].width = max_len
                
            # Mengaplikasikan border, posisi teks ke tengah, dan menebalkan Header
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if cell.row == 1: # Jika ini baris pertama (Header)
                        cell.font = Font(bold=True)
                        
        excel_data = output.getvalue()
        
        # 3. Menampilkan tombol unduh
        st.sidebar.download_button(
            label="Unduh Laporan",
            data=excel_data,
            file_name=f"Laporan_KMP_{start_date}_sd_{end_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        
        show_preview = st.sidebar.toggle("Tampilkan Visual Laporan")
    else:
        st.sidebar.warning("Tidak ada riwayat pada tanggal tersebut.")

st.sidebar.markdown("---")

# === PENGATURAN GARIS PINTU ===
st.sidebar.subheader("Pengaturan Garis Pintu")
idx_orient = 0 if cfg["LINE_ORIENT"] == "Vertikal" else 1
LINE_ORIENT = st.sidebar.selectbox("Orientasi Garis Pintu", ["Vertikal", "Horizontal"], index=idx_orient)

if LINE_ORIENT == "Vertikal":
    LINE_POS = st.sidebar.slider("Posisi Koordinat (X)", 0, 640, cfg["LINE_POS"])
    idx_dir = 0 if cfg["ENTRY_DIR"] == "Kiri ke Kanan" else 1
    ENTRY_DIR = st.sidebar.radio("Definisi Arah 'Masuk':", ["Kiri ke Kanan", "Kanan ke Kiri"], index=idx_dir)
else:
    LINE_POS = st.sidebar.slider("Posisi Koordinat (Y)", 0, 480, cfg["LINE_POS"])
    idx_dir = 0 if cfg["ENTRY_DIR"] == "Atas ke Bawah" else 1
    ENTRY_DIR = st.sidebar.radio("Definisi Arah 'Masuk':", ["Atas ke Bawah", "Bawah ke Atas"], index=idx_dir)

# === KONFIGURASI ZONA & WAKTU ===
with st.sidebar.expander("Konfigurasi Kotak & Waktu Tunggu", expanded=False):
    st.markdown("**Durasi Konversi (Detik)**")
    STAFF_LIMIT = st.slider("Waktu Tunggu Staf", 1, 60, cfg["STAFF_LIMIT"]) 
    BUYER_LIMIT = st.slider("Waktu Tunggu Pembeli", 1, 60, cfg["BUYER_LIMIT"]) 

    st.markdown("---")
    st.markdown("**Zona Staf (Biru)**")
    stf_x = st.slider("Staf: Posisi X", 0, 640, cfg["stf_x"])
    stf_y = st.slider("Staf: Posisi Y", 0, 480, cfg["stf_y"])
    stf_w = st.slider("Staf: Lebar", 50, 640, cfg["stf_w"])
    stf_h = st.slider("Staf: Tinggi", 50, 480, cfg["stf_h"])
    STAFF_ZONE = np.array([[stf_x, stf_y], [stf_x+stf_w, stf_y], [stf_x+stf_w, stf_y+stf_h], [stf_x, stf_y+stf_h]], np.int32)
    
    st.markdown("**Zona Kasir (Oranye)**")
    ksr_x = st.slider("Kasir: Posisi X", 0, 640, cfg["ksr_x"])
    ksr_y = st.slider("Kasir: Posisi Y", 0, 480, cfg["ksr_y"])
    ksr_w = st.slider("Kasir: Lebar", 50, 640, cfg["ksr_w"])
    ksr_h = st.slider("Kasir: Tinggi", 50, 480, cfg["ksr_h"])
    CASHIER_ZONE = np.array([[ksr_x, ksr_y], [ksr_x+ksr_w, ksr_y], [ksr_x+ksr_w, ksr_y+ksr_h], [ksr_x, ksr_y+ksr_h]], np.int32)

new_cfg = {
    "LINE_ORIENT": LINE_ORIENT, "LINE_POS": LINE_POS, "ENTRY_DIR": ENTRY_DIR,
    "STAFF_LIMIT": STAFF_LIMIT, "BUYER_LIMIT": BUYER_LIMIT,
    "stf_x": stf_x, "stf_y": stf_y, "stf_w": stf_w, "stf_h": stf_h,
    "ksr_x": ksr_x, "ksr_y": ksr_y, "ksr_w": ksr_w, "ksr_h": ksr_h
}
save_config(new_cfg)

@st.cache_resource
def load_model():
    return YOLO('best.pt')
model = load_model()

# ==========================================
# RENDER UTAMA BERDASARKAN MODE (PREVIEW / DASHBOARD)
# ==========================================
if show_preview and not df_export.empty:
    # MODE 1: TAMPILAN PRATINJAU LAPORAN
    st.markdown('<div class="metric-card"><div style="color:white; font-weight:600; font-size:20px; margin-bottom:10px;">📊 Pratinjau Laporan (Visualisasi Ekspor CSV)</div>', unsafe_allow_html=True)
    st.dataframe(df_export, use_container_width=True, hide_index=True)
    
    colA, colB = st.columns(2)
    with colA:
        # Chart 1: Perbandingan Pengunjung vs Pembeli
        fig_comp = go.Figure(data=[
            go.Bar(name='Pengunjung', x=df_export['Tanggal'].astype(str), y=df_export['Total Pengunjung'], marker_color='#00C9A7'),
            go.Bar(name='Pembeli', x=df_export['Tanggal'].astype(str), y=df_export['Total Pembeli'], marker_color='#D9568B')
        ])
        fig_comp.update_layout(title="Perbandingan Pengunjung vs Pembeli", barmode='group', paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font_color='#A0AEC0', hoverlabel=dict(bgcolor="#1E293B", font_color="white"))
        fig_comp.update_yaxes(showgrid=True, gridcolor='rgba(30, 77, 140, 0.3)', gridwidth=1)
        st.plotly_chart(fig_comp, use_container_width=True)
        
    with colB:
        # Chart 2: Menyesuaikan apakah 1 hari atau rentang hari
        if len(df_export) > 1:
            # Jika rentang hari: Tampilkan Hari Paling Ramai vs Sepi
            df_sorted = df_export.sort_values(by='Total Pengunjung', ascending=True)
            fig_ext = go.Figure(go.Bar(
                x=df_sorted['Total Pengunjung'], y=df_sorted['Tanggal'].astype(str), orientation='h', marker_color='#82B1FF'
            ))
            fig_ext.update_layout(title="Peringkat Hari (Paling Sepi ke Ramai)", paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font_color='#A0AEC0', hoverlabel=dict(bgcolor="#1E293B", font_color="white"))
            fig_ext.update_xaxes(showgrid=True, gridcolor='rgba(30, 77, 140, 0.3)', gridwidth=1)
            st.plotly_chart(fig_ext, use_container_width=True)
        else:
            # Jika hanya 1 hari: Tampilkan Gauge Tingkat Konversi
            rate = df_export.iloc[0]['Tingkat Konversi (%)']
            fig_gauge = go.Figure(go.Indicator(
                mode = "gauge+number",
                value = rate,
                title = {'text': "Tingkat Konversi", 'font': {'color': '#A0AEC0'}},
                gauge = {'axis': {'range': [0, 100], 'tickcolor': '#A0AEC0'},
                         'bar': {'color': "#D9568B"}}
            ))
            fig_gauge.update_layout(height=350, paper_bgcolor='rgba(0,0,0,0)', font_color='#A0AEC0')
            st.plotly_chart(fig_gauge, use_container_width=True)
            
    st.markdown('</div>', unsafe_allow_html=True)

else:
    # MODE 2: DASHBOARD KAMERA UTAMA (Default)
    st.markdown('<div style="color:white; font-weight:600; font-size:16px; margin-bottom:10px;">📹 Tayangan Langsung</div>', unsafe_allow_html=True)
    FRAME_WINDOW = st.empty()
    st.markdown("<br>", unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns(4)
    ph_in = col1.empty()
    ph_buy = col2.empty()
    ph_rate = col3.empty()
    ph_occ = col4.empty()

    def update_metrics_ui():
        c_in = st.session_state.count_in
        c_buy = st.session_state.count_buyer
        c_out = st.session_state.count_out
        rate = round((c_buy / c_in * 100), 1) if c_in > 0 else 0
        occ = max(0, c_in - c_out) 
        
        ph_in.markdown(f'<div class="metric-card"><div class="metric-title">TOTAL MASUK HARI INI</div><div class="metric-value val-white">{c_in}</div><div class="metric-sub">Pembaruan langsung</div></div>', unsafe_allow_html=True)
        ph_buy.markdown(f'<div class="metric-card"><div class="metric-title">TOTAL PEMBELI</div><div class="metric-value val-pink">{c_buy}</div><div class="metric-sub">Waktu tunggu terkonfirmasi</div></div>', unsafe_allow_html=True)
        ph_rate.markdown(f'<div class="metric-card"><div class="metric-title">TINGKAT KONVERSI</div><div class="metric-value val-teal">{rate}%</div><div class="metric-sub">Pengunjung → Pembeli</div></div>', unsafe_allow_html=True)
        ph_occ.markdown(f'<div class="metric-card"><div class="metric-title">PENGUNJUNG DI DALAM</div><div class="metric-value val-white">{occ}</div><div class="metric-sub">Di dalam toko saat ini</div></div>', unsafe_allow_html=True)

    update_metrics_ui()

    df_h, df_d = get_chart_data()
    col_left, col_right = st.columns([6, 4])

    with col_left:
        st.markdown('<div style="color:white; font-weight:600; margin-bottom:10px;">Lalu Lintas Per Jam Hari Ini</div>', unsafe_allow_html=True)
        if not df_h.empty:
            fig1 = go.Figure()
            fig1.add_trace(go.Bar(
                x=df_h['jam'].astype(str) + ':00', 
                y=df_h['total'], 
                marker_color='#D9568B', 
                name='Pengunjung',
                hovertemplate='<b>WAKTU %{x}</b><br>TOTAL %{y}<extra></extra>'
            ))
            fig1.update_layout(
                height=280, paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', font_color='#A0AEC0', 
                margin=dict(l=0, r=0, t=10, b=0), bargap=0.05,
                hoverlabel=dict(bgcolor="#1E293B", font_size=12, font_family="Inter", font_color="white")
            )
            fig1.update_xaxes(showgrid=False, linecolor='#1E4D8C')
            fig1.update_yaxes(showgrid=True, gridcolor='rgba(30, 77, 140, 0.3)', gridwidth=1)
            st.plotly_chart(fig1, use_container_width=True)
        else:
            st.info("Belum ada data jam ini.")

        with st.expander("Log Aktivitas Terkini", expanded=False):
            LOG_WINDOW = st.empty()

    with col_right:
        st.markdown('<div style="color:white; font-weight:600; margin-bottom:10px;">Pengunjung vs Pembeli (7 Hari)</div>', unsafe_allow_html=True)
        if not df_d.empty:
            fig2 = go.Figure(data=[
                go.Bar(name='Pengunjung', x=df_d['tanggal'], y=df_d['Pengunjung'], marker_color='#00C9A7', marker_line_width=0),
                go.Bar(name='Pembeli', x=df_d['tanggal'], y=df_d['Pembeli'], marker_color='#D9568B', marker_line_width=0)
            ])
            fig2.update_layout(
                height=560, barmode='group', paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)', 
                font_color='#A0AEC0', margin=dict(l=0, r=0, t=10, b=0), showlegend=False,
                hoverlabel=dict(bgcolor="#1E293B", font_size=12, font_family="Inter", font_color="white")
            )
            fig2.update_xaxes(showgrid=False, linecolor='#1E4D8C')
            fig2.update_yaxes(showgrid=True, gridcolor='rgba(30, 77, 140, 0.3)', gridwidth=1)
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Data historis tidak ditemukan.")

    # --- LOOP UTAMA KAMERA ---
    if run_camera and video_path is not None:
        cap = cv2.VideoCapture(video_path)
        if video_source == "Kamera Live":
            cap.set(3, 640); cap.set(4, 480)
            
        frame_count = 0
        
        while run_camera:
            ret, frame = cap.read()
            if not ret:
                st.success("✅ Pemutaran video selesai!")
                break
            
            frame_count += 1
            if frame_count % 2 == 0: continue
            
            h_asli, w_asli = frame.shape[:2]
            target_w, target_h = 640, 480
            scale = min(target_w / w_asli, target_h / h_asli)
            new_w, new_h = int(w_asli * scale), int(h_asli * scale)
            frame_resized = cv2.resize(frame, (new_w, new_h))
            canvas = np.zeros((target_h, target_w, 3), dtype=np.uint8)
            x_offset = (target_w - new_w) // 2
            y_offset = (target_h - new_h) // 2
            canvas[y_offset:y_offset+new_h, x_offset:x_offset+new_w] = frame_resized
            frame = canvas

            if LINE_ORIENT == "Vertikal":
                cv2.line(frame, (LINE_POS, 0), (LINE_POS, 480), (30, 77, 140), 2)
            else:
                cv2.line(frame, (0, LINE_POS), (640, LINE_POS), (30, 77, 140), 2)

            cv2.polylines(frame, [STAFF_ZONE], True, (200, 100, 50), 2)
            cv2.putText(frame, "ZONA STAF", (STAFF_ZONE[0][0], STAFF_ZONE[0][1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (200, 100, 50), 1)

            cv2.polylines(frame, [CASHIER_ZONE], True, (53, 107, 255), 2)
            cv2.putText(frame, "KASIR", (CASHIER_ZONE[0][0], CASHIER_ZONE[0][1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (53, 107, 255), 1)

            results = model.track(frame_kecil, persist=True, tracker="bytetrack.yaml", device=0, imgsz=480)
            
            if results[0].boxes is not None and results[0].boxes.id is not None:
                boxes = results[0].boxes.xyxy.cpu().numpy()
                ids = results[0].boxes.id.int().cpu().numpy()

                for box, track_id in zip(boxes, ids):
                    x1, y1, x2, y2 = map(int, box)
                    cx, cy = (x1+x2)//2, (y1+y2)//2
                    
                    is_staff = track_id in st.session_state.staff_ids
                    is_buyer = track_id in st.session_state.buyer_ids
                    
                    if cv2.pointPolygonTest(STAFF_ZONE, (cx, cy), False) >= 0 and not is_staff and not is_buyer:
                        if track_id not in st.session_state.staff_zone_timers:
                            st.session_state.staff_zone_timers[track_id] = time.time()
                        else:
                            elapsed = time.time() - st.session_state.staff_zone_timers[track_id]
                            if elapsed >= STAFF_LIMIT:
                                st.session_state.staff_ids.add(track_id)
                                log_to_database(track_id, "Staf Aktif")
                                add_log(track_id, "Staf Terdeteksi", "Zona Staf", f"{int(elapsed)}s")
                                is_staff = True
                    elif track_id in st.session_state.staff_zone_timers:
                        del st.session_state.staff_zone_timers[track_id]

                    if not is_staff:
                        if cv2.pointPolygonTest(CASHIER_ZONE, (cx, cy), False) >= 0 and not is_buyer:
                            if track_id not in st.session_state.cashier_zone_timers:
                                st.session_state.cashier_zone_timers[track_id] = time.time()
                            else:
                                elapsed = time.time() - st.session_state.cashier_zone_timers[track_id]
                                if elapsed >= BUYER_LIMIT:
                                    st.session_state.buyer_ids.add(track_id)
                                    st.session_state.count_buyer += 1
                                    log_to_database(track_id, "Pembeli Baru")
                                    add_log(track_id, "Pembeli Baru", "Kasir", f"{int(elapsed)}s")
                        elif track_id in st.session_state.cashier_zone_timers:
                            del st.session_state.cashier_zone_timers[track_id]

                        if LINE_ORIENT == "Vertikal":
                            if track_id not in st.session_state.track_states:
                                st.session_state.track_states[track_id] = 'kiri' if cx < LINE_POS else 'kanan'
                            else:
                                current_state = st.session_state.track_states[track_id]
                                
                                if current_state == 'kiri' and cx > LINE_POS + 15:
                                    if ENTRY_DIR == "Kiri ke Kanan":
                                        st.session_state.count_in += 1
                                        log_to_database(track_id, "Masuk")
                                        add_log(track_id, "Masuk", "Pintu Utama")
                                    else:
                                        st.session_state.count_out += 1
                                        log_to_database(track_id, "Keluar")
                                        add_log(track_id, "Keluar", "Pintu Utama")
                                    st.session_state.track_states[track_id] = 'kanan'
                                    
                                elif current_state == 'kanan' and cx < LINE_POS - 15:
                                    if ENTRY_DIR == "Kiri ke Kanan":
                                        st.session_state.count_out += 1
                                        log_to_database(track_id, "Keluar")
                                        add_log(track_id, "Keluar", "Pintu Utama")
                                    else:
                                        st.session_state.count_in += 1
                                        log_to_database(track_id, "Masuk")
                                        add_log(track_id, "Masuk", "Pintu Utama")
                                    st.session_state.track_states[track_id] = 'kiri'
                        else:
                            if track_id not in st.session_state.track_states:
                                st.session_state.track_states[track_id] = 'atas' if cy < LINE_POS else 'bawah'
                            else:
                                current_state = st.session_state.track_states[track_id]
                                
                                if current_state == 'atas' and cy > LINE_POS + 15:
                                    if ENTRY_DIR == "Atas ke Bawah":
                                        st.session_state.count_in += 1
                                        log_to_database(track_id, "Masuk")
                                        add_log(track_id, "Masuk", "Pintu Utama")
                                    else:
                                        st.session_state.count_out += 1
                                        log_to_database(track_id, "Keluar")
                                        add_log(track_id, "Keluar", "Pintu Utama")
                                    st.session_state.track_states[track_id] = 'bawah'
                                    
                                elif current_state == 'bawah' and cy < LINE_POS - 15:
                                    if ENTRY_DIR == "Atas ke Bawah":
                                        st.session_state.count_out += 1
                                        log_to_database(track_id, "Keluar")
                                        add_log(track_id, "Keluar", "Pintu Utama")
                                    else:
                                        st.session_state.count_in += 1
                                        log_to_database(track_id, "Masuk")
                                        add_log(track_id, "Masuk", "Pintu Utama")
                                    st.session_state.track_states[track_id] = 'atas'

                    color = (53, 107, 255) if is_buyer else ((200, 100, 50) if is_staff else (167, 201, 0))
                    label = "Pembeli" if is_buyer else ("Staf" if is_staff else "Pengunjung")
                    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 1)
                    cv2.rectangle(frame, (x1, y1-20), (x1+80, y1), color, -1)
                    cv2.putText(frame, f"ID:{track_id} {label}", (x1+5, y1-5), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0,0,0), 1)

            if frame_count % 5 == 0 or frame_count == 1:
                update_metrics_ui()
                LOG_WINDOW.markdown(render_log_table(), unsafe_allow_html=True)

            FRAME_WINDOW.image(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB), use_container_width=True)
        cap.release()
    else:
        with col_left:
            LOG_WINDOW.markdown(render_log_table(), unsafe_allow_html=True)